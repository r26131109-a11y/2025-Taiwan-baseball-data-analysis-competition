# -*- coding: utf-8 -*-
import os
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager

# --- 設定 ---
teams = [
    "光芒", "藍鳥", "金鶯", "紅襪", "洋基",
    "守護者", "皇家", "老虎", "雙城", "白襪",
    "太空人", "運動家", "遊騎兵", "天使", "水手",
    "勇士", "大都會", "國民", "費城人", "馬林魚",
    "海盜", "釀酒人", "紅雀", "小熊", "紅人",
    "響尾蛇", "道奇", "落磯", "教士", "巨人"
]

base_url = "https://tw.sports.yahoo.com/mlb/teams/{team}/schedule/?season=2025&scheduleType=list"
output_file = r"您的電腦位置"

# --- 初始化 Chrome ---
options = Options()
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

all_data = []

for team in teams:
    url = base_url.format(team=team)
    print(f"抓取 {team} 賽程中... {url}")
    driver.get(url)
    time.sleep(1)

    try:
        rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")

        for row in rows:
            text = row.text.strip()
            if not text:
                continue

            parts = text.split()
            if len(parts) < 5:
                continue

            date = parts[0]
            home_away = parts[1]
            opponent = parts[2]
            result = parts[3]
            score = " ".join(parts[4:])

            all_data.append({
                "球隊": team,
                "日期": date,
                "主/客": home_away,
                "對手": opponent,
                "勝敗": result,
                "比分": score
            })

    except Exception as e:
        print(f"{team} 抓取失敗：{e}")
        continue

driver.quit()

df = pd.DataFrame(all_data)
print(f"共抓取 {len(df)} 場比賽資料")

if not df.empty:
    if not os.path.exists(output_file):
        with pd.ExcelWriter(output_file, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, index=False, sheet_name="工作表1")
        print(f"已建立新檔案 {output_file}")
    else:
        from openpyxl import load_workbook
        book = load_workbook(output_file)
        if "工作表1" not in book.sheetnames:
            with pd.ExcelWriter(output_file, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, index=False, sheet_name="工作表1")
        else:
            with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                startrow = writer.sheets["工作表1"].max_row
                df.to_excel(writer, index=False, sheet_name="工作表1", header=False, startrow=startrow)
        print(f"已成功寫入 {output_file}")
else:
    print("沒有抓到任何資料！")